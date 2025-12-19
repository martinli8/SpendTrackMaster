import pandas as pd
import numpy as np
from datetime import datetime, date
import calendar
from typing import Dict, List, Any, Optional, Tuple

def get_month_name(month_num: int) -> str:
    """Get month name from month number"""
    return calendar.month_name[month_num]

def calculate_prorated_amount(amount: float, frequency: str) -> float:
    """Calculate monthly prorated amount based on frequency"""
    frequency_multipliers = {
        'monthly': 1.0,
        'quarterly': 1.0 / 3.0,
        'semi-annually': 1.0 / 6.0,
        'annually': 1.0 / 12.0
    }
    
    return amount * frequency_multipliers.get(frequency, 1.0)

def parse_bank_csv(file_content: str, filename: str) -> List[Dict]:
    """Parse bank CSV file and return list of transaction dictionaries"""
    try:
        # Read CSV content
        from io import StringIO
        df = pd.read_csv(StringIO(file_content))
        
        # Standardize column names (remove spaces and make lowercase)
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
        
        transactions = []
        for _, row in df.iterrows():
            # Skip empty rows
            if pd.isna(row.get('description', '')) or row.get('description', '').strip() == '':
                continue
            
            # Parse transaction date
            try:
                if 'transaction_date' in df.columns:
                    trans_date = pd.to_datetime(row['transaction_date']).date()
                elif 'date' in df.columns:
                    trans_date = pd.to_datetime(row['date']).date()
                else:
                    # Use first date column found
                    date_cols = [col for col in df.columns if 'date' in col]
                    if date_cols:
                        trans_date = pd.to_datetime(row[date_cols[0]]).date()
                    else:
                        trans_date = datetime.now().date()
            except:
                trans_date = datetime.now().date()
            
            # Parse post date if available
            post_date = None
            if 'post_date' in df.columns and not pd.isna(row['post_date']):
                try:
                    post_date = pd.to_datetime(row['post_date']).date()
                except:
                    post_date = None
            
            # Parse amount
            try:
                amount = float(row.get('amount', 0))
            except:
                amount = 0.0
            
            # Determine transaction type
            transaction_type = row.get('type', '').strip()
            if not transaction_type:
                # Infer type from amount
                if amount > 0:
                    transaction_type = 'Credit'
                else:
                    transaction_type = 'Debit'
            
            # Categorize based on description
            description = str(row.get('description', '')).strip()
            category = categorize_transaction(description)
            
            transaction = {
                'transaction_date': trans_date,
                'post_date': post_date,
                'description': description,
                'category': category,
                'type': transaction_type,
                'amount': amount,
                'memo': str(row.get('memo', '')).strip(),
                'source_file': filename
            }
            
            transactions.append(transaction)
        
        return transactions
    
    except Exception as e:
        raise ValueError(f"Error parsing CSV file: {str(e)}")

def parse_bank_excel(file_content: bytes, filename: str) -> List[Dict]:
    """Parse bank Excel file and return list of transaction dictionaries. Supports standard format and American Express statements."""
    try:
        from io import BytesIO
        import openpyxl
        
        # First, try to detect if this is an American Express statement
        # Read raw sheet to check format
        wb = openpyxl.load_workbook(BytesIO(file_content))
        ws = wb.active
        
        # Check if row 7 contains Amex-style headers (Date, Receipt, Description, Amount)
        is_amex_format = False
        try:
            row7_values = [cell.value for cell in ws[7]]
            if len(row7_values) >= 4 and row7_values[0] == 'Date' and 'Description' in str(row7_values):
                is_amex_format = True
        except:
            pass
        
        # Parse based on detected format
        if is_amex_format:
            transactions = _parse_amex_excel(ws, filename)
        else:
            # Standard format parsing
            df = pd.read_excel(BytesIO(file_content), engine='openpyxl')
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
            transactions = _parse_standard_excel(df, filename)
        
        return transactions
    
    except Exception as e:
        raise ValueError(f"Error parsing Excel file: {str(e)}")

def _parse_amex_excel(ws, filename: str) -> List[Dict]:
    """Parse American Express statement format Excel file"""
    transactions = []
    
    # Map Amex categories to user's categories
    category_mapping = {
        'airline': 'Travel',
        'hotel': 'Travel',
        'travel': 'Travel',
        'mobile telecom': 'Utilities',
        'internet services': 'Utilities',
        'education': 'Business School',
        'merchandise & supplies': 'Household Goods',
        'groceries': 'Groceries',
        'restaurants': 'Eating out',
        'gas stations': 'Gas',
        'health & medical': 'Health',
        'entertainment': 'Fun / Misc',
        'gifts': 'Gifts',
        'fees & adjustments': 'Bills'
    }
    
    # Start from row 8 (after headers in row 7)
    for row_idx, row in enumerate(ws.iter_rows(min_row=8, values_only=False), start=8):
        try:
            # Extract values from columns
            date_cell = row[0].value
            description = str(row[2].value).strip() if row[2].value else ''
            amount = row[3].value
            amex_category = str(row[11].value).strip() if row[11].value else ''
            
            # Skip if no description
            if not description or description == '':
                continue
            
            # Parse date
            try:
                trans_date = pd.to_datetime(date_cell).date()
            except:
                continue
            
            # Parse amount
            try:
                amount = float(amount) if amount else 0.0
            except:
                continue
            
            # Map Amex category to user's category, or auto-categorize
            category = 'Uncategorized'
            if amex_category:
                # Find matching category in mapping
                for amex_key, user_cat in category_mapping.items():
                    if amex_key.lower() in amex_category.lower():
                        category = user_cat
                        break
            
            # If still uncategorized, use description-based categorization
            if category == 'Uncategorized':
                category = categorize_transaction(description)
            
            # Determine transaction type based on amount
            transaction_type = 'Debit' if amount < 0 else 'Credit'
            
            transaction = {
                'transaction_date': trans_date,
                'post_date': None,
                'description': description,
                'category': category,
                'type': transaction_type,
                'amount': amount,
                'memo': amex_category,
                'source_file': filename
            }
            
            transactions.append(transaction)
        
        except Exception as e:
            continue
    
    return transactions

def _parse_standard_excel(df: pd.DataFrame, filename: str) -> List[Dict]:
    """Parse standard bank statement Excel format"""
    transactions = []
    
    for _, row in df.iterrows():
        # Skip empty rows
        if pd.isna(row.get('description', '')) or row.get('description', '').strip() == '':
            continue
        
        # Parse transaction date
        try:
            if 'transaction_date' in df.columns:
                trans_date = pd.to_datetime(row['transaction_date']).date()
            elif 'date' in df.columns:
                trans_date = pd.to_datetime(row['date']).date()
            else:
                # Use first date column found
                date_cols = [col for col in df.columns if 'date' in col]
                if date_cols:
                    trans_date = pd.to_datetime(row[date_cols[0]]).date()
                else:
                    trans_date = datetime.now().date()
        except:
            trans_date = datetime.now().date()
        
        # Parse post date if available
        post_date = None
        if 'post_date' in df.columns and not pd.isna(row['post_date']):
            try:
                post_date = pd.to_datetime(row['post_date']).date()
            except:
                post_date = None
        
        # Parse amount
        try:
            amount = float(row.get('amount', 0))
        except:
            amount = 0.0
        
        # Determine transaction type
        transaction_type = row.get('type', '').strip()
        if not transaction_type:
            if amount > 0:
                transaction_type = 'Credit'
            else:
                transaction_type = 'Debit'
        
        # Categorize based on description
        description = str(row.get('description', '')).strip()
        category = categorize_transaction(description)
        
        transaction = {
            'transaction_date': trans_date,
            'post_date': post_date,
            'description': description,
            'category': category,
            'type': transaction_type,
            'amount': amount,
            'memo': str(row.get('memo', '')).strip(),
            'source_file': filename
        }
        
        transactions.append(transaction)
    
    return transactions

def categorize_transaction(description: str) -> str:
    """Auto-categorize transaction based on description"""
    description_lower = description.lower()
    
    # Grocery stores
    grocery_keywords = [
        'whole foods', 'trader joe', 'safeway', 'kroger', 'walmart grocery', 'target grocery',
        'costco', 'sam\'s club', 'grocery', 'market', 'fresh', 'organic', 'food lion', 'harris teeter'
    ]
    
    # Eating out
    eating_out_keywords = [
        'restaurant', 'cafe', 'coffee', 'starbucks', 'mcdonald\'s', 'chipotle',
        'burger', 'pizza', 'taco', 'dining', 'kitchen', 'doordash', 'uber eats',
        'grubhub', 'takeout', 'delivery'
    ]
    
    # Household goods
    household_keywords = [
        'amazon', 'target', 'walmart', 'best buy', 'home depot', 'lowes', 
        'bed bath', 'ikea', 'costco', 'household', 'furniture', 'appliance',
        'cleaning', 'supplies'
    ]
    
    # Gas
    gas_keywords = [
        'shell', 'exxon', 'chevron', 'bp', 'mobil', 'sunoco', 'gas station',
        'fuel', 'gasoline', 'petrol'
    ]
    
    # Utilities
    utility_keywords = [
        'electric', 'gas company', 'water', 'internet', 'phone',
        'cable', 'utility', 'power', 'energy', 'verizon', 'comcast',
        'at&t', 'spectrum'
    ]
    
    # Health
    health_keywords = [
        'cvs', 'walgreens', 'pharmacy', 'doctor', 'medical', 'hospital',
        'dentist', 'health', 'prescription', 'medicine', 'clinic'
    ]
    
    # Travel
    travel_keywords = [
        'airline', 'flight', 'hotel', 'airbnb', 'uber', 'lyft', 'taxi',
        'rental car', 'airport', 'booking', 'expedia', 'travel'
    ]
    
    # Bills
    bills_keywords = [
        'insurance', 'loan', 'credit card', 'mortgage', 'rent',
        'subscription', 'membership', 'payment', 'autopay'
    ]
    
    # Fun/Misc
    fun_keywords = [
        'netflix', 'hulu', 'spotify', 'apple music', 'cinema',
        'theater', 'movie', 'concert', 'game', 'entertainment',
        'amazon prime', 'youtube', 'disney'
    ]
    
    # Gifts
    gift_keywords = [
        'gift', 'present', 'flowers', 'hallmark', 'card'
    ]
    
    # Check each category
    for keyword in grocery_keywords:
        if keyword in description_lower:
            return 'Groceries'
    
    for keyword in eating_out_keywords:
        if keyword in description_lower:
            return 'Eating out'
    
    for keyword in household_keywords:
        if keyword in description_lower:
            return 'Household Goods'
    
    for keyword in gas_keywords:
        if keyword in description_lower:
            return 'Gas'
    
    for keyword in utility_keywords:
        if keyword in description_lower:
            return 'Utilities'
    
    for keyword in health_keywords:
        if keyword in description_lower:
            return 'Health'
    
    for keyword in travel_keywords:
        if keyword in description_lower:
            return 'Travel'
    
    for keyword in bills_keywords:
        if keyword in description_lower:
            return 'Bills'
    
    for keyword in fun_keywords:
        if keyword in description_lower:
            return 'Fun / Misc'
    
    for keyword in gift_keywords:
        if keyword in description_lower:
            return 'Gifts'
    
    # Default category
    return 'Uncategorized'

def validate_file_format(filename: str) -> bool:
    """Validate if file format is supported"""
    supported_extensions = ['.csv', '.xlsx', '.xls']
    return any(filename.lower().endswith(ext) for ext in supported_extensions)

def format_currency(amount: float) -> str:
    """Format amount as currency string"""
    return f"${abs(amount):,.2f}"

def calculate_month_difference(start_date: date, end_date: date) -> int:
    """Calculate number of months between two dates"""
    return (end_date.year - start_date.year) * 12 + (end_date.month - start_date.month)

def get_date_range_months(start_date: date, end_date: date) -> List[Tuple[int, int]]:
    """Get list of (year, month) tuples for date range"""
    months = []
    current_date = start_date.replace(day=1)
    
    while current_date <= end_date:
        months.append((current_date.year, current_date.month))
        
        # Move to next month
        if current_date.month == 12:
            current_date = current_date.replace(year=current_date.year + 1, month=1)
        else:
            current_date = current_date.replace(month=current_date.month + 1)
    
    return months

def clean_amount_string(amount_str: str) -> float:
    """Clean amount string and convert to float"""
    if pd.isna(amount_str):
        return 0.0
    
    # Remove currency symbols and commas
    cleaned = str(amount_str).replace('$', '').replace(',', '').strip()
    
    # Handle parentheses for negative amounts
    if cleaned.startswith('(') and cleaned.endswith(')'):
        cleaned = '-' + cleaned[1:-1]
    
    try:
        return float(cleaned)
    except ValueError:
        return 0.0
